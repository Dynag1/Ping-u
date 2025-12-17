# This Python file uses the following encoding: utf-8

import src.var as var
import os

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    load_workbook = None

try:
    from PySide6.QtCore import QModelIndex
    from PySide6.QtWidgets import QFileDialog, QMessageBox
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    class QModelIndex: pass
    class QFileDialog:
        @staticmethod
        def getSaveFileName(*args, **kwargs): return (None, None)
        @staticmethod
        def getOpenFileName(*args, **kwargs): return (None, None)
    class QMessageBox:
        @staticmethod
        def information(*args): pass
        @staticmethod
        def critical(*args): pass
        @staticmethod
        def warning(*args): pass
        Ok = 1

name = ""


def chSave(self):
    if not GUI_AVAILABLE:
        # En mode headless, ces fonctions ne sont pas supportées ou doivent être gérées différemment
        return None

    filename, _ = QFileDialog.getSaveFileName(
        parent=self,
        caption=self.tr("Enregistrer le fichier"),
        dir=os.getcwd() + os.sep + "bd",
        filter=self.tr("Fichiers XLS (*.xlsx);;Tous fichiers (*.*)")
    )

    if not filename:
        return

    if not filename:
        return
    return filename


def chOpen(self):
    if not GUI_AVAILABLE:
        return None

    filename, _ = QFileDialog.getOpenFileName(
        parent=self,
        caption=self.tr("Ouvrir un fichier"),
        dir=os.getcwd() + os.sep + "bd",
        filter=self.tr("Fichiers xlsx (*.xlsx);;Tous fichiers (*.*)")
    )

    if not filename:
        return
    return filename


def saveExcel(self, tree_model):
    if not GUI_AVAILABLE:
        return
        
    if not OPENPYXL_AVAILABLE:
        QMessageBox.critical(self, self.tr("Erreur"), self.tr("Le module openpyxl n'est pas installé. Impossible d'exporter en Excel."), QMessageBox.Ok)
        return

    try:
        name = chSave(self)  # À remplacer par ta méthode de sélection de fichier
        if not name:
            return

        # Créer le classeur Excel
        workbook = Workbook()
        sheet = workbook.active
        # Entêtes avec Site
        headers = [self.tr("IP"), self.tr("Nom"), self.tr("Mac"), self.tr("Port"), self.tr("Latence"), self.tr("Site")]
        sheet.append(headers)
        # Parcourir le modèle
        for row in range(tree_model.rowCount()):
            # Récupérer les index des colonnes
            ip_index = tree_model.index(row, 1)
            nom_index = tree_model.index(row, 2)
            mac_index = tree_model.index(row, 3)
            port_index = tree_model.index(row, 4)
            latence_index = tree_model.index(row, 5)
            site_index = tree_model.index(row, 8)  # Colonne Site (était Comm)
            # Récupérer les données
            ip = ip_index.data()
            nom = nom_index.data()
            mac = mac_index.data()
            port = port_index.data()
            latence = latence_index.data()
            site = site_index.data()

            # Écrire dans Excel
            sheet.append([
                str(ip) if ip else "",
                str(nom) if nom else "",
                str(mac) if mac else "",
                str(port) if port else "",
                str(latence) if latence else "",
                str(site) if site else ""
            ])

        workbook.save(filename=f"{name}")
        QMessageBox.information(
            self,
            self.tr("Succès"),
            self.tr("Données sauvegardés"),
            QMessageBox.Ok
        )
    except Exception as e:
        print(e)

def openExcel(self, tree_model):
    if not GUI_AVAILABLE:
        return
        
    if not OPENPYXL_AVAILABLE:
        QMessageBox.critical(self, self.tr("Erreur"), self.tr("Le module openpyxl n'est pas installé. Impossible d'importer depuis Excel."), QMessageBox.Ok)
        return

    filename = chOpen(self)
    if not filename:
        return
        
    tree_model.removeRows(0, tree_model.rowCount())

    # Charger le workbook
    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            ip = str(row[0]) if row[0] else ""
            nom = str(row[1]) if len(row) > 1 and row[1] else ip
            mac = str(row[2]) if len(row) > 2 and row[2] else ""
            port = str(row[3]) if len(row) > 3 and row[3] else ""
            latence = ""
            site = str(row[5]) if len(row) > 5 and row[5] else ""  # Colonne Site

            # Vérifier si l'IP existe déjà
            ip_exists = False
            for i in range(tree_model.rowCount()):
                existing_ip = tree_model.index(i, 1).data()  # Colonne 1 = IP
                if existing_ip == ip:
                    msg = self.tr("L'adresse existe déjà")
                    QMessageBox.warning(self, self.tr("Doublon"), f"{msg} : {ip}")
                    ip_exists = True
                    break

            if not ip_exists and ip:
                # Ajouter au modèle
                parent = QModelIndex()  # Racine
                tree_model.insertRow(tree_model.rowCount(parent), parent)
                new_row = tree_model.rowCount(parent) - 1

                tree_model.setData(tree_model.index(new_row, 1, parent), ip)
                tree_model.setData(tree_model.index(new_row, 2, parent), nom)
                tree_model.setData(tree_model.index(new_row, 3, parent), mac)
                tree_model.setData(tree_model.index(new_row, 4, parent), port)
                tree_model.setData(tree_model.index(new_row, 5, parent), latence)
                tree_model.setData(tree_model.index(new_row, 8, parent), site)  # Colonne Site
        QMessageBox.information(
            self,
            self.tr("Succès"),
            self.tr("Données importés"),
            QMessageBox.Ok
        )
    except Exception as e:
        msg = self.tr("Erreur lors de la lecture")
        QMessageBox.critical(self, self.tr("Erreur"), f"{msg} : {str(e)}")


def export_xls_web(tree_model, filepath):
    """
    Export XLS pour l'API web (sans GUI)
    Format: IP, Nom, Mac, Site
    """
    if not OPENPYXL_AVAILABLE:
        raise Exception("Le module openpyxl n'est pas installé")
    
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hôtes"
        
        # En-têtes
        headers = ["IP", "Nom", "Mac", "Site"]
        sheet.append(headers)
        
        # Style pour les en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row in range(tree_model.rowCount()):
            ip = tree_model.item(row, 1).text() if tree_model.item(row, 1) else ""
            nom = tree_model.item(row, 2).text() if tree_model.item(row, 2) else ""
            mac = tree_model.item(row, 3).text() if tree_model.item(row, 3) else ""
            site = tree_model.item(row, 8).text() if tree_model.item(row, 8) else ""  # Colonne Site
            
            sheet.append([ip, nom, mac, site])
        
        # Ajuster la largeur des colonnes
        sheet.column_dimensions['A'].width = 18  # IP
        sheet.column_dimensions['B'].width = 30  # Nom
        sheet.column_dimensions['C'].width = 20  # Mac
        sheet.column_dimensions['D'].width = 20  # Site
        
        workbook.save(filepath)
        return True
        
    except Exception as e:
        raise Exception(f"Erreur export XLS: {str(e)}")


def import_xls_web(tree_model, filepath):
    """
    Import XLS pour l'API web (sans GUI)
    Format attendu: IP, Nom, Mac, Site (en-têtes sur la première ligne)
    """
    if not OPENPYXL_AVAILABLE:
        raise Exception("Le module openpyxl n'est pas installé")
    
    try:
        workbook = load_workbook(filename=filepath)
        sheet = workbook.active
        
        # Import QStandardItem selon le mode
        if GUI_AVAILABLE:
            from PySide6.QtGui import QStandardItem
        else:
            class QStandardItem:
                def __init__(self, text=""): 
                    self._text = str(text) if text else ""
                def text(self): return self._text
                def setText(self, text): self._text = str(text)
        
        imported_count = 0
        duplicates = 0
        
        # Collecter les IPs existantes
        existing_ips = set()
        for i in range(tree_model.rowCount()):
            item = tree_model.item(i, 1)
            if item:
                existing_ips.add(item.text())
        
        # Parcourir les lignes (ignorer l'en-tête)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
                
            ip = str(row[0]).strip() if row[0] else ""
            nom = str(row[1]).strip() if len(row) > 1 and row[1] else ip
            mac = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            site = str(row[3]).strip() if len(row) > 3 and row[3] else ""  # Colonne Site
            
            if not ip:
                continue
            
            # Vérifier les doublons
            if ip in existing_ips:
                duplicates += 1
                continue
            
            # Ajouter au modèle
            items = [
                QStandardItem(str(tree_model.rowCount())),  # 0: Id
                QStandardItem(ip),                          # 1: IP
                QStandardItem(nom),                         # 2: Nom
                QStandardItem(mac),                         # 3: Mac
                QStandardItem(""),                          # 4: Port
                QStandardItem(""),                          # 5: Latence
                QStandardItem(""),                          # 6: Temp
                QStandardItem(""),                          # 7: Suivi
                QStandardItem(site),                        # 8: Site
                QStandardItem("")                           # 9: Excl
            ]
            
            tree_model.appendRow(items)
            existing_ips.add(ip)
            imported_count += 1
        
        return {
            'imported': imported_count,
            'duplicates': duplicates
        }
        
    except Exception as e:
        raise Exception(f"Erreur import XLS: {str(e)}")